import json
from collections import defaultdict
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.mixins import PermissionRequiredMixin, LoginRequiredMixin
from django.contrib.sessions.exceptions import SessionInterrupted
from django.core.exceptions import ValidationError
from django.db import transaction, models
from django.db.models import Sum, Count, Q, Prefetch, OuterRef, Subquery
from django.db.models.functions.datetime import TruncMonth
from django.http import HttpResponseForbidden
from django.shortcuts import redirect
from django.contrib.auth import login, logout
from django.urls import reverse
from django.urls.base import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.http import require_http_methods
from django.views.generic.base import TemplateView, View
from django.views.generic.edit import UpdateView, CreateView, FormView
from openpyxl.styles import Font, Alignment

from gestecole import settings
from gestecole.utils.calculatrice_bulletin import calculer_moyennes_coefficients, calculer_moyenne_generale
from gestecole.utils.decorateurs import administrateur_required, requires_approval, \
    censeur_only, comptable_only, directeur_required, censeur_required, surveillant_required, comptable_required, \
    multi_role_required, tenant_required
from gestecole.utils.file_handlers import save_files_to_temp
from gestecole.utils.idgenerateurs import IDGenerator
from gestecole.utils.messageries import SmsOrangeService
from gestecole.utils.paiements import MalitelMoneyAPI, OrangeMoneyAPI
from gestecole.utils.services import MyLogin  # , LoginAdminFrom, get_client_ip
from . import forms
from .forms import LoginForm, LoginFormAgent, LoginAdminFrom, EleveCreationForm, BulletinForm, ExamenForm, \
    EnseignantCreationForm, MatiereForm, ClasseForm, PeriodeForm, AbsenceForm, PeriodePaiementForm, \
    EmploiDuTempsForm, PaiementForm, ValiderPaiementForm, PaiementEspeceForm, HistoriqueAcademiqueForm
from django.conf import settings
from datetime import datetime

from django_ratelimit.decorators import ratelimit

from .models import Note, Administrateur, Eleve, Matiere, Classe, Enseignant, \
    BulletinPerformance, BulletinMatiere, NoteExamen, Examen, Periode, logger, HistoriqueAcademique, PeriodePaiement, Paiement

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from django.http import HttpResponse

from .services.periodique.file_genereale import get_active_period, FiltreService
from weasyprint import HTML
from .tasks import envoyer_notifications_initiales, verifier_paiements_echelonnes

ul = MyLogin()


@tenant_required
@ratelimit(key='post:identifiant', rate='5/h', method='POST')
@csrf_protect
@ratelimit(key='ip', rate='10/m', block=True)
def Admininstrateur(request):
    if request.method == settings.METHODE_POST:
        form = LoginAdminFrom(data=request.POST)
        if form.is_valid():
            identifiant = form.cleaned_data['identifiant']
            password = form.cleaned_data['password']
            user = ul.login_user_Admin(identifiant, password)
            print(user)
            if user is not None:
                if user.is_admin and user.is_authenticated:
                    try:
                        login(request, user, backend='APP_G2S.auth_backends.AdminBackend')
                        request.session.set_expiry(settings.SESSION_COOKIE_AGE)  # Session de 1h
                        request.session['last_login'] = str(timezone.now())
                        return redirect('dashboard_admin')
                    except Exception as e:
                        messages.error(request, f"Erreur de connexion : {str(e)}")
                else:
                    messages.error(request, "Vous n'avez pas les droits de Super Agent.")
            else:
                messages.error(request, "Identifiant ou mot de passe incorrect.")
        else:
            messages.error(request, "Formulaire invalide.")
    else:
        form = LoginAdminFrom()
    return render(request, 'APP_G2S/connexion_admin.html', {'form': form})



@tenant_required
@ratelimit(key='post:telephone', rate='5/h', method='POST')
@csrf_protect
def login_view_eleve(request):
    if request.method == settings.METHODE_POST:
        form = LoginForm(data=request.POST)
        if form.is_valid():
            telephone = form.cleaned_data.get('telephone', None)
            password = form.cleaned_data.get('password', None)
            user = ul.login_user(telephone, password)
            if user is None:
                '''
                je vais gere la session ici par la cockie ou par le stokage de request au local
                '''
                messages.error(request, f'Telephone {telephone} inconnu')
                return redirect('login')
            login(request, user, backend='APP_G2S.auth_backends.TelephoneBackend')
        if request.user.is_authenticated:
            print(request.user.telephone)
            return redirect('dashboard_admin')
    else:
        form = LoginForm()
    request.session.set_expiry(settings.SESSION_COOKIE_AGE)  # Session de 1h
    request.session['last_login'] = str(timezone.now())
    return render(request, 'APP_G2S/connexions/connexion.html', {'forms': form})


@tenant_required
@ratelimit(key='post:matricule', rate='5/h', method='POST')
@csrf_protect
def login_view_agent(request):
    if request.method == settings.METHODE_POST:
        form = LoginFormAgent(request.POST)
        if form.is_valid():
            matricule = form.cleaned_data['matricule']
            password = form.cleaned_data['password']
            user = ul.login_user_Agent(matricule, password)

            if user is None:
                messages.error(request, "Matricule incorrect ou agent inactif.")
                return redirect('login_agent')
            if hasattr(user, 'is_enseignant') and user.is_enseignant:
                login(request, user, backend='APP_G2S.auth_backends.MatriculeBackend')
                if request.user.is_authenticated:
                    request.session.set_expiry(settings.SESSION_COOKIE_AGE)  # Session de 1h
                    request.session['last_login'] = str(timezone.now())
                    return redirect('agent_dashboard')
            else:
                messages.error(request, "Mot de passe incorrect.")
        else:
            messages.error(request, "Données invalides.")
    else:
        form = LoginFormAgent()
    return render(request, 'APP_G2S/agent/login_agent.html', {'form': form})



@tenant_required
@administrateur_required
# @permission_required('APP_G2S.view_all', raise_exception=True)
@multi_role_required(directeur_required, censeur_required, surveillant_required, comptable_required)
def dashboard_admin(request):
    if request.user.identifiant:
       if request.user.is_authenticated:
            admin = Administrateur.objects.get(identifiant=request.user.identifiant)
            return render(request, "APP_G2S/composant-admin/dashboard.html", {"user": admin})
       try:
           user_data = request.session['user_data']
       except SessionInterrupted:
           return HttpResponse("Session expired or invalid.", user_data)
       return redirect('connexion_admin')
    return HttpResponse("FORBIDDEN")




@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required, surveillant_required, comptable_required)
def dashboard_eleve(request):

    return render(request, "APP_G2S/composant-admin/dashboard_eleve.html")

@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required, surveillant_required)
def liste_eleves(request):

    active_period = get_active_period()
    if not active_period:
        return render(request, 'APP_G2S/composant-admin/liste_eleves.html', {'Noperiode': None})
    matieres = Matiere.objects.all()

    # Récupération des paramètres de filtrage
    annee = request.GET.get('annee')
    classe_id = request.GET.get('classe')

    export_url = reverse('pdf_eleves') + f'?classe_id={classe_id}&annee={annee}'

    # Filtrage des classes selon l'année
    classes = FiltreService.get_classes(annee)
    eleves = Eleve.objects.filter(classe__in=classes)

    # Filtre supplémentaire par classe spécifique
    if classe_id and classe_id.isdigit():
        eleves = eleves.filter(classe__id=int(classe_id))

    context = {
        "eleves": eleves,
        "annees": FiltreService.get_academie_years(),
        "selected_annee": annee,
        "classes_dispo": classes,  # Remplace classes_uniques
        "selected_classe": classe_id,
        'matiere': matieres,
        'active_period': active_period,
        'export_url': export_url,
        'Noperiode': 1
    }
    return render(request, 'APP_G2S/composant-admin/liste_eleves.html', context)

@tenant_required
@multi_role_required(administrateur_required, censeur_required, surveillant_required, directeur_required)
def detail_eleve(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)
    notes = Note.objects.filter(eleve=eleve).select_related('matiere')
    note_examen = NoteExamen.objects.filter(eleve=eleve).select_related('matiere')
    absences = Absence.objects.filter(eleve=eleve).select_related('matiere')
    emploi = EmploiDuTemps.objects.filter(classe=eleve.classe).prefetch_related('matiere')

    print(eleve.profile_picture)
    context = {
        'eleve': eleve,
        'notes': notes,
        'notes_examen': note_examen,
        'emploi': emploi,
        'absences': absences,
    }
    return render(request, 'APP_G2S/composant-admin/detail_eleve.html', context)



@tenant_required
@multi_role_required(administrateur_required, censeur_required, directeur_required, surveillant_required)
def detail_enseignant(request, enseignant_id):
    active_period = get_active_period()
    enseignant = get_object_or_404(Enseignant, id=enseignant_id)

    context = {
        'enseignant': enseignant,
        'active_period': active_period
    }

    return render(request, 'APP_G2S/composant-admin/detail_enseignant.html', context)


@tenant_required
@multi_role_required(administrateur_required, censeur_required, surveillant_required, directeur_required)
def emploi_du_temps(request):
    active_period = get_active_period()
    classe_selected = request.GET.get('classe')

    queryset = EmploiDuTemps.objects.all().select_related(
        'classe', 'matiere', 'enseignant'
    ).order_by('date', 'start_time')  # Utiliser les champs réels

    if classe_selected:
        queryset = queryset.filter(classe__id=classe_selected)

    context = {
        'emploi': queryset,
        'classes': Classe.objects.all(),
        'classe_selected': classe_selected,
        'active_period': active_period
    }
    return render(request, 'APP_G2S/composant-admin/emploi_du_temps.html', context)


@tenant_required
@administrateur_required
@censeur_only
def liste_bulletins(request):
    active_period = get_active_period()
    if not active_period:
        return render(request, 'APP_G2S/composant-admin/liste_bulletins.html', {'Noperiode': None})

    annee = request.GET.get('annee')
    classe_id = request.GET.get('classe')
    periode_id = request.GET.get('periode')

    # Récupération des classes et périodes
    classes = FiltreService.get_classes(annee).filter(periode=active_period)  # Ajout du filtre par période active
    all_periodes = Periode.objects.filter(is_active=True).order_by('-annee_scolaire', '-numero')  # Filtrer les périodes actives

    # Définir la période sélectionnée
    selected_period = active_period
    if periode_id and periode_id.isdigit():
        try:
            selected_period = Periode.objects.get(id=int(periode_id))
        except Periode.DoesNotExist:
            messages.error(request, "La période sélectionnée n'existe pas ou est clôturée.")

    # Filtrage des élèves
    eleves = Eleve.objects.filter(classe__in=classes)
    if classe_id and classe_id.isdigit():
        eleves = eleves.filter(classe__id=int(classe_id))

    # Filtrage des bulletins
    bulletins = BulletinPerformance.objects.filter(
        periode=selected_period,
        eleve__in=eleves
    ).select_related('eleve').order_by('-date_creation')

    context = {
        'bulletins': bulletins,
        'active_period': selected_period,
        'all_periodes': all_periodes,
        "annees": FiltreService.get_academie_years(),
        "selected_annee": annee,
        "classes_dispo": classes,
        "selected_classe": classe_id,
        "selected_periode": periode_id,
        'Noperiode': 1
    }
    return render(request, 'APP_G2S/composant-admin/liste_bulletins.html', context)

@tenant_required
@administrateur_required
# @transaction.atomic
@censeur_only
def ajouter_bulletin(request):
    active_period = get_active_period() # la periode active
    # Initialisation des variables


    eleve = None
    classe = None
    periode_active = None
    form = BulletinForm(request.POST or None)
    context = {
        'form': form,
        'eleve': None,
        'eleves_classe': [],
        'notes_existantes': [],
        'notes_existantes_E': [],
        'matieres_classe': [],
        'moyennes_coefficient': {},
        'moyenne_generale': None,
        'classement': None,
        'periode_active': None,
        'bulletin_existant': None,
        'examens_valides': True,
        'active_period': active_period
    }

    try:
        # Récupération de la période active
        periode_active = Periode.objects.prefetch_related('classe').get(is_active=True)
        context['periode_active'] = periode_active
    except Periode.DoesNotExist:
        messages.error(request, "Aucune période scolaire active configurée")
        return render(request, 'APP_G2S/composant-admin/ajouter_bulletin.html', context)

    # Gestion des paramètres GET/POST
    classe_id = request.GET.get('classe') or request.POST.get('classe')
    eleve_id = request.GET.get('eleve') or request.POST.get('eleve')

    if classe_id:
        try:
            classe = Classe.objects.prefetch_related('eleves', 'matieres').get(id=classe_id)
        except Classe.DoesNotExist:
            messages.error(request, "Classe introuvable")

        # Récupération de l'élève et mise à jour de la classe si nécessaire
    if eleve_id:
        try:
            eleve = Eleve.objects.select_related('classe').get(id=eleve_id)
            context['eleve'] = eleve

            # Définir la classe à partir de l'élève si non spécifiée
            if not classe and eleve.classe:
                classe = eleve.classe
                context['classe_selectionnee'] = classe

            # Vérification de la cohérence classe/élève
            if classe and eleve.classe != classe:
                messages.error(request, "L'élève ne fait pas partie de cette classe")
                return redirect('ajouter_bulletin')

            # Récupération des données pédagogiques
            matieres_classe = eleve.classe.matieres.all()
            context['matieres_classe'] = matieres_classe

            # Optimisation des requêtes de notes
            notes_existantes = Note.objects.filter(
                eleve=eleve,
                periode=periode_active
            ).select_related('matiere')

            notes_existantes_E = NoteExamen.objects.filter(
                eleve=eleve,
                periode=periode_active
            ).select_related('matiere')

            # Conversion en dictionnaire pour accès rapide
            notes_classe_dict = {n.matiere_id: n.valeur for n in notes_existantes}
            notes_exam_dict = {n.matiere_id: n.note for n in notes_existantes_E}

            # Calcul des moyennes
            moyennes_coefficient = calculer_moyennes_coefficients(
                notes_classe_dict,
                notes_exam_dict,
                matieres_classe
            )

            context.update({
                'notes_existantes': notes_existantes,
                'notes_existantes_E': notes_existantes_E,
                'moyennes_coefficient': moyennes_coefficient,
                'moyenne_generale': calculer_moyenne_generale(moyennes_coefficient, matieres_classe),
                'bulletin_existant': BulletinPerformance.objects.filter(
                    eleve=eleve,
                    periode=periode_active
                ).first()
            })

        except Eleve.DoesNotExist:
            messages.error(request, "Élève introuvable")

    # Gestion de la soumission du formulaire
    if request.method == 'POST':
        # Validation préalable
        if periode_active.cloture:
            messages.error(request, "Période clôturée - modifications impossibles")
            return redirect('ajouter_bulletin')

        # Vérification des examens associés
        examens_invalides = Examen.objects.filter(
            Q(periode=periode_active) &
            (Q(validite='FIN') | Q(date__lt=timezone.now().date()))
        ).exists()

        if examens_invalides:
            messages.error(request, "Examens invalides détectés")
            return redirect('ajouter_bulletin')

        if form.is_valid():
            try:
                # Création/mise à jour du bulletin
                bulletin, created = BulletinPerformance.objects.update_or_create(
                    eleve=eleve,
                    periode=periode_active,
                    defaults={
                        'appreciation': form.cleaned_data['appreciation'],
                        'moyenne_generale': context['moyenne_generale'],
                        'classes': eleve.classe,
                        'date_modification': timezone.now()
                    }
                )

                # Traitement des notes par matière
                bulletin_matiere_data = []
                for matiere in context['matieres_classe']:
                    note_classe = float(request.POST.get(f'note_classe_{matiere.id}', 0))
                    note_exam = float(request.POST.get(f'note_exam_{matiere.id}', 0))

                    # Calcul selon la pondération (classe: 1/3, examen: 2/3)
                    note_ponderee = (note_classe + (note_exam * 2)) / 3

                    bulletin_matiere_data.append(BulletinMatiere(
                        bulletin=bulletin,
                        matiere=matiere,
                        note=note_ponderee
                    ))

                # Mise à jour en masse
                BulletinMatiere.objects.bulk_create(
                    bulletin_matiere_data,
                    update_conflicts=True,
                    update_fields=['note'],
                    unique_fields=['bulletin', 'matiere']
                )

                messages.success(request,
                                 f"Bulletin {'mis à jour' if not created else 'créé'} avec succès"
                                 )

                # Redirection avec paramètres
                redirect_params = []
                if classe: redirect_params.append(f'classe={classe.id}')
                if eleve: redirect_params.append(f'eleve={eleve.id}')
                return redirect(f"{reverse('ajouter_bulletin')}?{'&'.join(redirect_params)}")

            except Exception as e:
                logger.error(f"Erreur d'enregistrement: {str(e)}", exc_info=True)
                messages.error(request, f"Erreur critique lors de l'enregistrement: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

    # Préparation du contexte final
    context.update({
        'examens_valides': not Examen.objects.filter(
            periode=periode_active,
            validite='FIN'
        ).exists()
    })

    # Gestion des requêtes AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'moyenne_generale': context['moyenne_generale']
        })

    return render(request, 'APP_G2S/composant-admin/ajouter_bulletin.html', context)


@tenant_required
@administrateur_required
@censeur_only
def detail_bulletin(request, bulletin_id):
    bulletin = get_object_or_404(BulletinPerformance, id=bulletin_id)
    notes_details = bulletin.get_notes_details()

    class_stats = BulletinPerformance.objects.filter(
        eleve__classe=bulletin.eleve.classe,
        periode=bulletin.periode
    ).aggregate(
        avg_moyenne=Avg('moyenne_generale'),
        total=Count('id')
    )

    context = {
        'bulletin': bulletin,
        'notes_details': notes_details,
        'class_average': class_stats['avg_moyenne'] or 0,
        'class_total': class_stats['total'] or 0,
    }
    return render(request, 'APP_G2S/composant-admin/detail_bulletin.html', context)

@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required, surveillant_required)
def enseignant_gestion(request):
    enseignants = Enseignant.objects.all()
    context = {
        'enseignants': enseignants,
    }
    return render(request, 'APP_G2S/composant-admin/liste_enseignant.html', context)




@tenant_required
@administrateur_required
@censeur_only
def ajouter_enseignant(request):
    if request.method == settings.METHODE_POST:
        form = EnseignantCreationForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                enseignant = form.save(commit=False)

                # Génération de l'ID enseignant
                enseignant.identifiant = IDGenerator.generate_teacher_id()
                password = IDGenerator.generatriceMDP_default()

                from django.contrib.auth.hashers import make_password
                enseignant.password = make_password(password)

                enseignant.is_enseignant = True

                enseignant.save()
                form.save_m2m()
                sms_service = SmsOrangeService()
                message = (
                    f"Bienvenue {enseignant.nom_complet}!\n"
                    f"Identifiant: {enseignant.identifiant}\n"
                    f"Mot de passe temporaire: {password}\n"
                    f"Valide 15 minutes. A changer après connexion."
                )
                success, _ = sms_service.envoyer_sms_orange(enseignant.telephone, message)
                if success:
                    messages.success(request, "Enseignant créé avec succès ! SMS envoyé.")
                else:
                    messages.warning(request, "Enseignant créé mais échec d'envoi SMS")

                return redirect('enseignant_gestion')

            except Exception as e:
                messages.error(request, f"Erreur lors de la création : {str(e)}")
                print(e)
        else:
            messages.error(request, "Formulaire invalide. Veuillez corriger les erreurs.")
    else:
        form = EnseignantCreationForm()

    enseignants = Enseignant.objects.all()
    return render(request, 'APP_G2S/composant-admin/ajouter_enseignant.html', {
        'enseignants': enseignants,
        'form': form
    })

@tenant_required
@administrateur_required
@requires_approval(action_type='supprimer_enseignant', model=Enseignant, id_param='enseignant_id')
def supprimer_enseignant(request, enseignant_id):
    enseignant = get_object_or_404(Enseignant, id=enseignant_id)
    if request.method == 'POST':
        enseignant.delete()
        messages.success(request, "Enseignant supprimé avec succès.")
        return redirect('enseignant_gestion')
    return render(request, 'APP_G2S/composant-admin/supprimer_enseignant.html', {'enseignant': enseignant})


@tenant_required
@administrateur_required
@requires_approval(action_type='modifier_enseignant', model=Enseignant, id_param='enseignant_id')
def modifier_enseignant(request, enseignant_id):
    enseignant = get_object_or_404(Enseignant, id=enseignant_id)
    if request.method == 'POST':
        form = EnseignantCreationForm(request.POST, request.FILES, instance=enseignant)
        if form.is_valid():
            form.save()
            messages.success(request, "Enseignant modifié avec succès !")
            return redirect('enseignant_gestion')
    else:
        form = EnseignantCreationForm(instance=enseignant)
    return render(request, 'APP_G2S/composant-admin/modifier_enseignant.html', {'form': form})


@tenant_required
@administrateur_required
@censeur_only
def ajouter_eleve(request):
    if request.method == 'POST':
        form = EleveCreationForm(request.POST, request.FILES)
        try:
            if form.is_valid():
                try:
                    eleve = form.save(commit=False)
                    eleve.is_eleve = True

                    password = IDGenerator.generatriceMDP_default()

                    from django.contrib.auth.hashers import make_password
                    eleve.password = make_password(password)
                    user_id = f"{eleve.nom}-{eleve.prenom}"
                    with transaction.atomic():
                        try:
                            temp_files = save_files_to_temp(request.FILES, user_id)
                            request.session['temp_files'] = temp_files
                        except Exception as e:
                            print(f"Erreur au niveau temp_files: {e}")
                            messages.error(request, f"Erreur au niveau temp_files: {e}")
                            return redirect('eleve_gestion')

                        sms_service = SmsOrangeService()
                        message = (
                            f"Bienvenue {eleve.prenom} {eleve.nom}!\n"
                            f"Identifiant: {eleve.identifiant}\n"
                            f"Mot de passe temporaire: {password}\n"
                            f"Valide 15 minutes. A changer après connexion."
                        )
                        success, _ = sms_service.envoyer_sms_orange(eleve.telephone, message)
                        if success:
                            eleve.save()
                            form.save_m2m()
                            messages.success(request, "Élève créé avec succès. SMS envoyé!")
                            return redirect('eleve_gestion')
                        else:
                            messages.warning(request, "Élève créé mais échec d'envoi SMS")
                        return redirect('eleve_gestion')
                except ValidationError as e:
                    messages.error(request, str(e))
                except Exception as e:
                    messages.error(request, f"Erreur système : {str(e)}")
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"Erreur sur le champ {field}: {error}")
                messages.error(request, "Formulaire invalide")
        except Exception as e:
            messages.error(request, f"Erreur Vous n'avez pas ajoute image ou image trop eleve, max 5MB. Si persite contacter ce numero +223 94 30 63 02")
    else:
        form = EleveCreationForm()

    return render(request, 'APP_G2S/composant-admin/ajouter_eleve.html', {'form': form})


@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required, surveillant_required)
def ajouter_emploi(request):
    if request.method == 'POST':
        form = EmploiDuTempsForm(request.POST)
        try:
            if form.is_valid():
                form.save()
                messages.success(request, "Cours ajouté avec succès !")
                return redirect('ajouter_emploi')
            else:
                # Afficher les erreurs dans la console
                print("Erreurs du formulaire:", form.errors)
                # Afficher les erreurs aux utilisateurs via messages
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        except Exception as e:
            messages.error(request, f"ERREUR: {str(e)}")
            print(f"ERREUR: {str(e)}")
    else:
        form = EmploiDuTempsForm()

    emplois = EmploiDuTemps.objects.all().order_by('date', 'start_time')
    return render(request, 'APP_G2S/composant-admin/ajouter_emploi.html', {
        'form': form,
        'emplois': emplois
    })
@tenant_required
@administrateur_required
@censeur_only
def ajouter_note(request):

    return render(request, 'APP_G2S/composant-admin/ajouter_note.html')

@tenant_required
@administrateur_required
@censeur_only
def liste_examens(request):
    active_period = get_active_period()
    if not active_period:
        return render(request, 'APP_G2S/composant-admin/liste_examens.html', {'Noperiode': None})

    annee = request.GET.get('annee')
    classe_id = request.GET.get('classe')
    periode_id = request.GET.get('periode')

    # Récupération des classes et périodes
    classes = FiltreService.get_classes(annee).filter(periode=active_period)
    all_periodes = Periode.objects.filter(is_active=True).order_by('-annee_scolaire', '-numero')

    # Définir la période sélectionnée
    selected_period = active_period
    if periode_id and periode_id.isdigit():
        try:
            selected_period = Periode.objects.get(id=int(periode_id))
        except Periode.DoesNotExist:
            messages.error(request, "La période sélectionnée n'existe pas ou est clôturée.")

    # Filtrage des examens
    examens = Examen.objects.filter(
        periode=selected_period
    ).prefetch_related('classe').prefetch_related('matieres').order_by('-date')

    if classe_id and classe_id.isdigit():
        examens = examens.filter(classe__id=int(classe_id))

    context = {
        'examens': examens,
        'active_period': selected_period,
        'all_periodes': all_periodes,
        "annees": FiltreService.get_academie_years(),
        "selected_annee": annee,
        "classes_dispo": classes,
        "selected_classe": classe_id,
        "selected_periode": periode_id,
        'Noperiode': 1
    }
    return render(request, 'APP_G2S/composant-admin/liste_examens.html', context)


from django.db.models import Avg, Max, Min
from django.shortcuts import get_object_or_404, render

@tenant_required
@censeur_only
@administrateur_required
def detail_examen(request, examen_id):
    # Récupération des paramètres de filtre
    selected_annee = request.GET.get('annee')
    selected_classe = request.GET.get('classe')
    print(selected_classe)
    print(selected_annee)

    # Récupération de l'examen avec préchargement optimisé
    examen = get_object_or_404(
        Examen.objects.prefetch_related(
            Prefetch('classe', queryset=Classe.objects.all()),
            Prefetch('matieres')
        ),
        pk=examen_id
    )

    # Filtrage des classes selon l'année scolaire
    classes_query = examen.classe.all()
    if selected_annee:
        classes_query = classes_query.filter(periode__annee_scolaire=selected_annee)
        print(classes_query)

    # Filtrage supplémentaire si une classe est sélectionnée
    if selected_classe:
        classes_query = classes_query.filter(id=selected_classe)
        print(classes_query)

    # Récupération des élèves concernés
    eleves = Eleve.objects.filter(classe__in=classes_query)
    print(eleves)

    # Récupération des notes avec filtrage
    notes_examen = NoteExamen.objects.filter(
        examen=examen,
        eleve__in=eleves
    ).select_related('eleve', 'matiere').order_by('eleve__nom')

    # Calcul des statistiques
    stats = notes_examen.aggregate(
        avg_note=Avg('note'),
        max_note=Max('note'),
        min_note=Min('note')
    )

    # Préparation des données groupées
    notes_par_eleve = defaultdict(lambda: {'eleve': None, 'notes': {}, 'moyenne': 0})
    for note in notes_examen:
        eleve_data = notes_par_eleve[note.eleve.id]
        eleve_data['eleve'] = note.eleve
        eleve_data['notes'][note.matiere.id] = note.note

    # Calcul des moyennes par élève
    for eleve_data in notes_par_eleve.values():
        notes = list(eleve_data['notes'].values())
        eleve_data['moyenne'] = sum(notes) / len(notes) if notes else 0

    context = {
        'examen': examen,
        'classes': classes_query,
        'matieres': examen.matieres.all(),
        'notes_par_eleve': sorted(notes_par_eleve.values(), key=lambda x: x['eleve'].nom),
        'moyenne_generale': stats['avg_note'] or 0,
        'note_max': stats['max_note'] or 0,
        'note_min': stats['min_note'] or 0,
        'total_eleves': eleves.count(),
        'annees': FiltreService.get_academie_years(),
        'selected_annee': selected_annee,
        'classes_dispo': examen.classe.all(),
        'selected_classe': selected_classe,
    }
    return render(request, 'APP_G2S/composant-admin/detail_examen.html', context)

@tenant_required
@administrateur_required
@csrf_protect
@directeur_required
@requires_approval(action_type='modifier_notes', model=(Note, NoteExamen), id_param=None)
def saisir_notes(request):
    if request.method == 'POST':
        classe_id = request.POST.get('classe')
        matiere_id = request.POST.get('matiere')
        examen_id = request.POST.get('examen')

        try:
            # Validation des données requises
            if not all([classe_id, matiere_id]):
                raise ValidationError("Classe et matière sont obligatoires")

            classe = Classe.objects.get(id=classe_id)
            matiere = Matiere.objects.get(id=matiere_id)
            examen = Examen.objects.get(id=examen_id) if examen_id else None
            today = timezone.now().date()

            # Validation de la période
            periode = Periode.objects.filter(
                classe=classe,
                date_debut__lte=today,
                date_fin__gte=today
            ).first()

            if not periode:
                raise ValidationError("Aucune période active pour cette classe")

            if periode.cloture or not periode.is_active:
                raise ValidationError("Période clôturée/inactive")

            # Validation des contraintes d'examen
            if not examen and any(k.startswith('note_classe_') for k in request.POST):
                raise ValidationError("Les notes de classe nécessitent un examen associé")

            if examen:
                if not (examen.date <= today <= examen.date_fin):
                    raise ValidationError("Hors période de l'examen ou date de début invalide")

                periode_examen = Examen.objects.filter(
                    validite="EN_COURS", periode=periode
                ).first()

                if not periode_examen:
                    raise ValidationError("Aucun examen en cours pour cette période")

            # Traitement des notes dans une transaction
            with transaction.atomic():
                for key, value in request.POST.items():
                    if key.startswith('note_classe_') and not examen:
                        raise ValidationError("Notes de classe sans examen associé")

                    if key.startswith('note_classe_') or key.startswith('note_examen_'):
                        eleve_id = key.split('_')[2]
                        eleve = Eleve.objects.get(id=eleve_id)

                        if key.startswith('note_classe_'):
                            Note.objects.update_or_create(
                                eleve=eleve,
                                matiere=matiere,
                                classe=classe,
                                periode=periode,
                                examen_reference=periode_examen,
                                defaults={'valeur': float(value), 'date': today}
                            )
                        elif key.startswith('note_examen_'):
                            NoteExamen.objects.update_or_create(
                                eleve=eleve,
                                matiere=matiere,
                                periode=periode,
                                examen=examen,
                                defaults={'note': float(value), 'date': today}
                            )

            messages.success(request, "Notes sauvegardées")
            return redirect(f"{reverse('saisir_notes')}?classe={classe_id}")

        except (Classe.DoesNotExist, Matiere.DoesNotExist, Eleve.DoesNotExist) as e:
            messages.error(request, f"Référence introuvable : {str(e)}")
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Erreur système : {str(e)}")

    # Gestion de la requête GET
    classe_id = request.GET.get('classe') or request.POST.get('classe')
    classes = Classe.objects.all()
    matieres = Matiere.objects.all()
    examens_en_cours = Examen.objects.filter(validite='EN_COURS')
    examens_fin = Examen.objects.filter(validite='FIN')

    # Sélection de la classe avec gestion d'erreur
    classe = None
    if classe_id:
        try:
            classe = Classe.objects.prefetch_related('matieres').get(id=classe_id)
            matieres = classe.matieres.all()
        except Classe.DoesNotExist:
            messages.error(request, "Classe sélectionnée introuvable")

    return render(request, 'APP_G2S/composant-admin/saisir_notes.html', {
        'classes': classes,
        'matieres': matieres,
        'examens_en_cours': examens_en_cours,
        'examens_fin': examens_fin,
        'classe_selectionnee': classe,
        'active_period': get_active_period()
    })

@tenant_required
@administrateur_required
@censeur_only
def liste_notes(request):
    active_period = get_active_period()
    if not active_period:
        return render(request, 'APP_G2S/composant-admin/liste_notes.html', {'Noperiode': None})
    notes_classe = Note.objects.filter(periode=active_period).select_related('eleve', 'matiere')
    notes_examen = NoteExamen.objects.filter(periode=active_period).select_related('eleve', 'examen')

    return render(request, 'APP_G2S/composant-admin/liste_notes.html', {
        'notes_classe': notes_classe,
        'notes_examen': notes_examen,
        'active_period': active_period,
        'Noperiode': 1
    })

from django.http import JsonResponse
from django.views.decorators.http import require_GET

@require_GET
@tenant_required
@administrateur_required
@censeur_only # optionnel
def api_eleves(request):
    classe_id = request.GET.get('classe')
    matiere_id = request.GET.get('matiere')
    examen_id = request.GET.get('examen')

    eleves = Eleve.objects.filter(classe__id=classe_id)
    data = []
    today = timezone.now().date()

    periode = Periode.objects.active().filter(
        classe__id=classe_id,
        # date_debut__lte=today,
        # date_fin__gte=today,
        cloture=False
    ).first()
    print('periode', periode)

    for eleve in eleves:
        note_classe = float(0)
        note_examen = float(0)

        try:
            examen = Examen.objects.get(id=examen_id) if examen_id else None
        except Examen.DoesNotExist:
            messages.error(request, "Examen n'existe pas ou non enregistré")
        except Exception as e:
            print('dans api_eleves', str(e))
            logger.error(f"Erreur dans api_eleves: {str(e)}")

        try:
            if examen_id:
                examen = Examen.objects.get(id=examen_id)
                classe = examen.classe.prefetch_related('matieres').get(id=classe_id)
                # matieres = classe.matieres.get(id=matiere_id)
                examen_valide = (
                        examen.validite == 'EN_COURS' and
                        examen.date <= today <= examen.date_fin and
                        examen.periode.is_active and
                        not examen.periode.cloture and
                        classe.matieres.filter(id=matiere_id).exists()
                )

                if examen_valide:
                    note = NoteExamen.objects.filter(
                        eleve=eleve,
                        examen=examen,
                        matiere_id=matiere_id
                    ).first()
                    note_examen = note.note if note else float(0)

                    n_classe = Note.objects.filter(
                        eleve=eleve,
                        matiere__id=matiere_id,
                        periode=periode,
                        examen_reference=examen
                    ).first()
                    note_classe = n_classe.valeur if n_classe else float(0)
        except Examen.DoesNotExist:
            messages.error(request, "Examen n'existe pas ou non enregistré")
        except AttributeError as e:
            # e = examen.classe.prefetch_related('matieres').filter(id=matiere_id).first()
            # print('oooo', examen.classe.prefetch_related('matieres').matieres.filter(id=matiere_id).first())
            # print(e.matieres.first())

            print('probleme d\'attribut veuillez contacter le service +223 94 30 63 02', str(e))
            messages.error(request, 'probleme d\'attribut veuillez contacter le service +223 94 30 63 02', str(e))

        data.append({
            'id': eleve.id,
            'nom_complet': f"{eleve.prenom} {eleve.nom}",
            'note_classe': float(note_classe),
            'note_examen': float(note_examen),
        })

        # print('la data c\'est dire les donnees', data)

    return JsonResponse(data, safe=False)


@require_GET
@tenant_required
@administrateur_required
@censeur_only
def api_examens(request, examen_id):
    try:
        examen = Examen.objects.get(id=examen_id)
        print('examen.periode.cloture', examen.periode.cloture)
        return JsonResponse({
            'date': examen.date.strftime("%d-%m-%y"),
            'date_fin': examen.date_fin.strftime("%d-%m-%y"),
            'periode_active': examen.periode.is_active,
            'periode_cloture': examen.periode.cloture,
            'validite': examen.validite
        })
    except Examen.DoesNotExist:
        return JsonResponse({'error': 'Examen non trouvé'}, status=404)


@tenant_required
@censeur_only
@requires_approval(action_type='creer_examen', model=Examen, id_param=None)
@administrateur_required
def creer_examen(request):
    if request.method == 'POST':
        form = ExamenForm(request.POST)
        if form.is_valid():
            try:
                # # Validation 1 : Vérifier les matières via form.cleaned_data
                # if not form.cleaned_data.get('matiere'):
                #     form.add_error('matiere', "Sélectionnez au moins une matière")
                #     raise ValidationError("Matière manquante")

                # Validation 2 : Date cohérente
                if form.cleaned_data['validite'] == 'FIN' and form.cleaned_data['date'] > timezone.now().date():
                    form.add_error('validite', "Un examen terminé ne peut pas avoir une date future")
                    raise ValidationError("Date incohérente")

                # Sauvegarde de l'objet principal
                examen = form.save()  # Sauvegarde directe (pas besoin de commit=False)

                # # Sauvegarde des relations ManyToMany
                # form.save_m2m()  # Optionnel ici car form.save() le fait déjà

                messages.success(request, "Examen créé avec succès !")
                return redirect('liste_examens')

            except ValidationError as e:
                messages.error(request, str(e))
        else:
            # Afficher les erreurs du formulaire
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ExamenForm()

    return render(request, 'APP_G2S/composant-admin/creer_examen.html', {'form': form})


@tenant_required
@administrateur_required
@directeur_required
# @requires_approval(action_type='modifier_periode', model=Periode, id_param='periode_id')
def periode_scolaire(request):
    active_period = None
    form = None
    periodes_actives = Periode.objects.all()
    for i in periodes_actives:
        print(i.classe)
        print(dir(i.classe))
        print(dir(i))

    try:
        # Récupération de la période active existante
        active_period = Periode.objects.filter(is_active=True).first()
        form = PeriodeForm(instance=active_period)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la période active : {str(e)}", exc_info=True)
        messages.error(request, "Erreur de configuration : impossible de récupérer la période active.")
        form = PeriodeForm()

    if request.method == "POST":
        form = PeriodeForm(request.POST, instance=active_period)
        try:
            if form.is_valid():
                periode = form.save(commit=False)
                periode.save()
                form.save_m2m()
                messages.success(request, "Période enregistrée avec succès")
                return redirect('gestion_periodes')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
                messages.error(request, "Formulaire invalide. Veuillez corriger les erreurs.")
        except Exception as e:
            logger.error(f"Erreur critique lors de l'enregistrement de la période : {str(e)}", exc_info=True)
            messages.error(request, f"Erreur lors de l'enregistrement : {str(e)}")

    return render(request, 'APP_G2S/composant-admin/periode_scolaire.html', {
        'form': form,
        'periodes_actives': periodes_actives
    })


@tenant_required
@censeur_only
@requires_approval(action_type='modifier_periode', model=Periode, id_param='periode_id')
def modifier_periode(request, periode_id):
    periode = get_object_or_404(Periode, id=periode_id)
    if request.method == 'POST':
        form = PeriodeForm(request.POST, instance=periode)

        # Si un commentaire est fourni, l'ajouter aux données de la requête
        commentaire = request.POST.get('commentaire')
        if commentaire and request.user.role == 'CENSEUR':
            # Stocker le commentaire dans la session pour qu'il soit disponible lors de la création de la demande d'approbation
            request.session['commentaire_modification'] = commentaire

        if form.is_valid():
            form.save()
            messages.success(request, "Période modifiée avec succès !")
            return redirect('gestion_periodes')
        else:
            messages.error(request, "Erreur dans le formulaire. Veuillez corriger les erreurs.")
    else:
        form = PeriodeForm(instance=periode)

    context = {
        'form': form,
        'periode': periode,
        'is_censeur': request.user.role == 'CENSEUR'
    }
    return render(request, 'APP_G2S/composant-admin/modifier_periode.html', context)


@tenant_required
@censeur_only
@requires_approval(action_type='supprimer_periode', model=Periode, id_param='periode_id')
def supprimer_periode(request, periode_id):
    periode = get_object_or_404(Periode, id=periode_id)
    if request.method == 'POST':
        # Si un commentaire est fourni, l'ajouter aux données de la requête
        commentaire = request.POST.get('commentaire')
        if commentaire and request.user.role == 'CENSEUR':
            # Stocker le commentaire dans la session pour qu'il soit disponible lors de la création de la demande d'approbation
            request.session['commentaire_suppression'] = commentaire

        periode.delete()
        messages.success(request, "Période supprimée avec succès")
        return redirect('gestion_periodes')
    return render(request, 'APP_G2S/composant-admin/supprimer_periode.html', {
        'periode': periode,
        'is_censeur': request.user.role == 'CENSEUR'
    })


@tenant_required
@administrateur_required
@censeur_only
def ajouter_classe(request):
    if request.method == 'POST':
        form = ClasseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Classe ajoutée avec succès !")
            return redirect('liste_classes')
    else:
        form = ClasseForm()

    context = {
        'form': form,
        'enseignants': Enseignant.objects.all()
    }
    return render(request, 'APP_G2S/composant-admin/ajouter_classe.html', context)


@tenant_required
@administrateur_required
@requires_approval(action_type='modifier_classe', model=Classe, id_param='classe_id')
def modifier_classe(request, classe_id):
    classe = get_object_or_404(Classe, id=classe_id)
    if request.method == 'POST':
        form = ClasseForm(request.POST, instance=classe)
        if form.is_valid():
            form.save()
            messages.success(request, "Classe modifiée avec succès !")
            return redirect('liste_classes')
        else:
            messages.error(request, "Erreur dans le formulaire. Veuillez corriger les erreurs.")
    else:
        form = ClasseForm(instance=classe)

    context = {
        'form': form,
        'classe': classe,
        'enseignants': Enseignant.objects.all()
    }
    return render(request, 'APP_G2S/composant-admin/modifier_classe.html', context)


@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required, surveillant_required)
def liste_classes(request):
    active_period = get_active_period()
    classes = Classe.objects.all().select_related('responsable')
    return render(request, 'APP_G2S/composant-admin/liste_classes.html', {
        'classes': classes,
        'active_period' : active_period
    })


@tenant_required
@administrateur_required
@censeur_only
def ajouter_matiere(request):
    if request.method == 'POST':
        form = MatiereForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Matière ajoutée avec succès !")
            return redirect('liste_matieres')
    else:
        form = MatiereForm()

    return render(request, 'APP_G2S/composant-admin/ajouter_matiere.html', {'form': form})

@tenant_required
@administrateur_required
@censeur_only
def liste_matieres(request):
    active_period = get_active_period()
    matieres = Matiere.objects.all().prefetch_related('classe_set')
    print(matieres)
    print()
    return render(request, 'APP_G2S/composant-admin/liste_matiere.html', {
        'matieres': matieres,
        'active_period' : active_period
    })

@require_GET
@tenant_required
@administrateur_required
@censeur_only
def api_matieres(request):
    classe_id = request.GET.get('classe_id')
    if classe_id:
        matieres = Matiere.objects.filter(classe__id=classe_id).values('id', 'nom')
        return JsonResponse(list(matieres), safe=False)
    return JsonResponse([], safe=False)

@tenant_required
@censeur_only
@requires_approval(action_type='modifier_notes_examen', model=NoteExamen, id_param=None)
def saisir_notes_examen(request, examen_id):
    examen = get_object_or_404(Examen, id=examen_id)
    eleves = Eleve.objects.filter(classe=examen.classe)

    if request.method == 'POST':
        for eleve in eleves:
            note = request.POST.get(f'note_{eleve.id}')
            try:
                note_value = float(note)
                if not (0 <= note_value <= 40):
                    messages.error(request, f"Note invalide pour {eleve.nom}")
                    continue
            except ValueError:
                messages.error(request, f"Valeur numérique invalide pour {eleve.nom}")
                continue
            if note:
                NoteExamen.objects.update_or_create(
                    eleve=eleve,
                    examen=examen,
                    defaults={'note': float(note)}
                )
        messages.success(request, "Notes enregistrées avec succès !")
        return redirect('detail_examen', examen_id=examen.id)

    return render(request, 'APP_G2S/composant-admin/saisir_notes.html', {
        'examen': examen,
        'eleves': eleves
    })


@tenant_required
@administrateur_required
def logout_view(request):
    logout(request)
    return redirect('connexion_admin')

from APP_G2S.tasks import *


@method_decorator(administrateur_required, name='dispatch')
@method_decorator(comptable_only, name='dispatch')
@method_decorator(transaction.atomic, name='post')
class CreerPeriodePaiementView(FormView):
    template_name = 'APP_G2S/composant-comptable/creer_periode_paiement.html'
    form_class = PeriodePaiementForm
    success_url = reverse_lazy('gestion_paiements')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'classes': Classe.objects.all(),
            'examens': Examen.objects.filter(validite='EN_COURS'),
            'montant_min': 5000
        })
        return context

    def form_valid(self, form):
        try:
            periode = form.save(commit=False)

            # Validation des dates
            if periode.date_debut >= periode.date_fin:
                form.add_error('date_fin', "La date de fin doit être postérieure à la date de début")
                return self.form_invalid(form)

            periode.save()
            form.save_m2m()  # Pour les relations ManyToMany (classes)

            # Planification des tâches asynchrones
            # envoyer_notifications_initiales.delay(periode.id)
            # planifier_rappels(periode.id)
            print("creer")
            messages.success(self.request,
                             f"Période {periode.nom} créée avec succès ! "
                             f"Notifications envoyées à {periode.classe.count()} classes."
                             )
            print("creer")
            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, f"Erreur critique : {str(e)}")
            logger.error(f"Erreur création période paiement : {str(e)}")
            return self.form_invalid(form)


@method_decorator(comptable_only, name='dispatch')
class GestionPaiementsView(TemplateView):
    template_name = 'APP_G2S/composant-comptable/gestion_paiements.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        periodes = PeriodePaiement.objects.all().prefetch_related('classe', 'examen')
        context.update({
            'periodes': periodes,
            'stats': {
                'total': periodes.count(),
                'actives': periodes.filter(date_fin__gte=timezone.now()).count(),
                'montant_moyen': periodes.aggregate(Avg('montant_total'))['montant_total__avg'] or 0
            }
        })
        return context





class ValiderPaiementView(PermissionRequiredMixin, UpdateView):
    permission_required = 'paiement.confirmer_paiement'
    model = Paiement
    form_class = ValiderPaiementForm
    template_name = 'APP_G2S/composant-comptable/validation_paiement.html'

    def form_valid(self, form):
        paiement = form.save(commit=False)
        if form.cleaned_data['statut_paiement'] == 'ANNULE':
            paiement._mettre_a_jour_tranche(reset=True)
        paiement.save()
        messages.success(self.request, "Statut du paiement mis à jour")
        return redirect('suivi-paiements')


class CaisseDashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'paiement.add_paiement'
    template_name = 'APP_G2S/composant-comptable/caisse_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['paiements_jour'] = Paiement.objects.filter(
            date_paiement__date=timezone.now().date(),
            caissier=self.request.user
        )
        return context

# @administrateur_required
@method_decorator(transaction.atomic, name='post')
class EncaissementView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'paiement.add_paiement'
    raise_exception = True
    permission_required = 'paiement.add_paiement'
    form_class = PaiementEspeceForm
    template_name = 'APP_G2S/composant-comptable/encaissement.html'
    success_url = reverse_lazy('caisse-dashboard')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['classes'] = Classe.objects.all().order_by('niveau', 'section')
        return context



    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        paiement = form.save(commit=False)
        paiement.caissier = self.request.user
        paiement.mode_paiement = 'ESPECES'
        paiement = form.save(commit=False)
        paiement.tranche = form.cleaned_data.get('tranche')
        paiement.save()

        # Mise à jour du statut de la tranche
        if paiement.tranche:
            self._update_tranche_status(paiement.tranche)

        # Génération PDF
        self._generate_quittance(paiement)

        messages.success(self.request, f"Paiement enregistré - Quittance #{paiement.numero_quittance}")
        return super().form_valid(form)

    def _update_tranche_status(self, tranche):
        total_paye = tranche.paiement_set.aggregate(total=Sum('montant_paye'))['total'] or 0

        if total_paye >= tranche.montant:
            tranche.statut = 'PAYE'
        elif total_paye > 0:
            tranche.statut = 'PARTIEL'
        tranche.save()

    def _generate_quittance(self, paiement):
        # Intégration de WeasyPrint pour générer PDF


        html_string = render_to_string('paiement/quittance_pdf.html', {'paiement': paiement})
        HTML(string=html_string).write_pdf(
            f"media/quittances/{paiement.numero_quittance}.pdf"
        )


class AnnulationPaiementView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'paiement.delete_paiement'
    model = Paiement
    fields = []
    template_name = 'APP_G2S/composant-comptable/annulation_paiement.html'
    success_url = reverse_lazy('caisse-dashboard')

    @transaction.atomic
    def form_valid(self, form):
        paiement = self.get_object()

        if paiement.est_confirme:
            messages.error(self.request, "Impossible d'annuler un paiement confirmé")
            return redirect('caisse-dashboard')

        # Remise à zéro de la tranche
        if paiement.tranche:
            tranche = paiement.tranche
            tranche.statut = 'ECHEU' if tranche.date_echeance < timezone.now().date() else 'NON_ECHEU'
            tranche.save()

        paiement.delete()
        messages.success(self.request, "Paiement annulé avec succès")
        return super().form_valid(form)


def get_tranche_info(request):
    periode_id = request.GET.get('periode_id')
    eleve_id = request.GET.get('eleve_id')

    periode = get_object_or_404(PeriodePaiement, id=periode_id)
    eleve = get_object_or_404(Eleve, id=eleve_id)

    tranche = periode.prochaine_tranche_eleve(eleve)

    return JsonResponse({
        'tranche_num': f"Tranche {tranche.ordre}" if tranche else "Complet",
        'montant_restant': float(tranche.montant_restant(eleve)) if tranche else 0
    })







@require_GET
@tenant_required
@administrateur_required
def charger_donnees_eleve(request, eleve_id):
    try:
        eleve = Eleve.objects.get(id=eleve_id)
    except Eleve.DoesNotExist:
        return JsonResponse({'error': 'Élève non trouvé'}, status=404)
    matieres = Matiere.objects.filter(classe=eleve.classe)

    matieres_data = []
    for matiere in matieres:
        note = Note.objects.filter(eleve=eleve, matiere=matiere).first()
        matieres_data.append({
            'id': matiere.id,
            'nom': matiere.nom,
            'coefficient': matiere.coefficient,
            'note_existante': note.valeur if note else None
        })

    return JsonResponse({
        'eleve': {
            'nom': eleve.nom,
            'prenom': eleve.prenom,
            'classe': str(eleve.classe),
            'telephone': str(eleve.telephone),
            'residence': eleve.residence
        },
        'matiere': matieres_data,
        # 'appreciation': eleve.appreciation_generale
    })


@require_GET
@tenant_required
@administrateur_required
def get_matieres_par_classe(request):
    eleve_id = request.GET.get('eleve_id')
    try:
        eleve = Eleve.objects.get(id=eleve_id)
        matieres = Matiere.objects.filter(classe=eleve.classe).distinct().values('id', 'nom', 'coefficient')
        return JsonResponse(list(matieres), safe=False)
    except Eleve.DoesNotExist:
        return JsonResponse([], safe=False)


@tenant_required
@administrateur_required
@requires_approval(action_type='modifier_statut_examen', model=Examen, id_param='examen_id')
def modifier_statut_examen(request, examen_id):
    examen = get_object_or_404(Examen, id=examen_id)

    if request.user.role not in ['DIRECTEUR', 'CENSEUR']:
        return HttpResponseForbidden("Permission refusée")

    if request.method == 'POST':
        nouveau_statut = request.POST.get('validite')
        if nouveau_statut in ['EN_COURS', 'FIN']:
            examen.validite = nouveau_statut
            examen.save()
            messages.success(request, "Statut de l'examen mis à jour")
            return redirect('liste_examens')

    return render(request, 'APP_G2S/composant-admin/modifier_statut_examen.html', {
        'examen': examen
    })


@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required, surveillant_required)
def liste_absences(request):
    active_period = get_active_period()
    absences = Absence.objects.select_related('eleve', 'emploi_du_temps').order_by('-date')

    # Filtres
    eleve_id = request.GET.get('eleve')
    date = request.GET.get('date')
    periode_id = request.GET.get('periode')
    classe_id = request.GET.get('classe')
    annee = request.GET.get('annee')


    if eleve_id:
        absences = absences.filter(eleve__id=eleve_id)
    if date:
        absences = absences.filter(date=date)
    if periode_id and periode_id.isdigit():
        absences = absences.filter(emploi_du_temps__periode__id=int(periode_id))
    if classe_id and classe_id.isdigit():
        absences = absences.filter(eleve__classe__id=int(classe_id))


    export_url = reverse('pdf_absences') + f'?classe_id={classe_id}&date_debut={date}&date_fin={date}'

    context = {
        'absences': absences,
        "annees": FiltreService.get_academie_years(),
        "selected_annee": annee,
        'eleves': Eleve.objects.all(),
        'periodes': Periode.objects.all(),
        'classes_dispo': Classe.objects.all(),
        'active_period': active_period,
        'selected_eleve': eleve_id,
        'selected_periode': periode_id,
        'selected_classe': classe_id,
        'selected_date': date,
        'export_url': export_url,
    }
    return render(request, 'APP_G2S/composant-admin/liste_absences.html', context)


@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required, surveillant_required)
def ajouter_absence(request):
    # Récupérer l'ID de la classe depuis les paramètres GET pour filtrer les emplois du temps
    classe_id = request.GET.get('classe_id')
    emplois = EmploiDuTemps.objects.all().order_by('date')
    eleves = Eleve.objects.all()
    if classe_id and classe_id.isdigit():
        emplois = emplois.filter(classe_id=classe_id)
        # emplois = emplois.filter(classe_id=classe_id)
        eleves = eleves.filter(classe_id=classe_id)

    if request.method == 'POST':
        form = AbsenceForm(request.POST)
        if form.is_valid():
            try:
                absence = form.save(commit=False)
                absence.date = absence.emploi_du_temps.date
                absence.save()
                # Ajout du nom et prénom de l'élève dans le message SMS
                message = (
                    f"Alerte absence : {absence.eleve.prenom} {absence.eleve.nom} absent(e) le {absence.date.strftime('%d/%m')} "
                    f"en {absence.emploi_du_temps.matiere.nom}."
                    f" Justification : {request.POST.get('url') or request.META.get('HTTP_REFERER') or request.build_absolute_uri()}/absences/{absence.id}"
                )
                sms_service = SmsOrangeService()
                sms_service.envoyer_sms_orange(
                    str(absence.eleve.telephone),
                    message
                )
                messages.success(request, "Absence enregistrée avec succès")
            except Exception as e:
                print(f"Erreur envoi SMS : {str(e)}")
                messages.error(request, f"Erreur : {str(e)}")
                logger.error(f"Erreur envoi SMS : {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = AbsenceForm()

    # Passer la liste filtrée des emplois du temps et la classe sélectionnée au template
    return render(request, 'APP_G2S/composant-admin/ajouter_absence.html', {
        'form': form,
        'emplois': emplois,
        'classes': Classe.objects.all(),
        'classe_selected': classe_id,
        'eleves': eleves
    })


@tenant_required
@administrateur_required
@requires_approval(action_type='supprimer_absence', model=Absence, id_param='absence_id')
def supprimer_absence(request, absence_id):
    absence = get_object_or_404(Absence, id=absence_id)
    if request.method == 'POST':
        absence.delete()
        messages.success(request, "Absence supprimée avec succès")
        return redirect('liste_absences')
    return render(request, 'APP_G2S/composant-admin/supprimer_absence.html', {'absence': absence})


@tenant_required
@administrateur_required
@requires_approval(action_type='modifier_absence', model=Absence, id_param='absence_id')
def modifier_absence(request, absence_id):
    absence = get_object_or_404(Absence, id=absence_id)
    if request.method == 'POST':
        form = AbsenceForm(request.POST, instance=absence)
        if form.is_valid():
            form.save()
            messages.success(request, "Absence modifiée avec succès")
            return redirect('liste_absences')
    else:
        form = AbsenceForm(instance=absence)
    return render(request, 'APP_G2S/composant-admin/modifier_absence.html', {'form': form})

def generer_recu_paiement(paiement):
    response = HttpResponse(content_type='application/pdf')
    filename = f"reçu_paiement_{paiement.id}_{paiement.date_paiement.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # En-tête de l'établissement
    story.append(Paragraph("ÉCOLE SECONDAIRE EXCELLENCE", styles['Title']))
    story.append(Paragraph("123 Avenue du Savoir, Ville", styles['Normal']))
    story.append(Paragraph("Tél: +223 00 00 00 00", styles['Normal']))
    story.append(Spacer(1, 12))

    # Contenu du reçu
    content = [
        f"Reçu N°: {paiement.id}",
        f"Date: {paiement.date_paiement.strftime('%d/%m/%Y %H:%M')}",
        f"Élève: {paiement.eleve}",
        f"Classe: {paiement.eleve.classe.nom}",
        f"Montant: {paiement.montant_paye} €",
        f"Mode de paiement: {paiement.get_mode_paiement_display()}",
        f"Encaissé par: {paiement.encaisser_par.get_full_name() if paiement.encaisser_par else 'Système'}"
    ]

    for line in content:
        story.append(Paragraph(line, styles['BodyText']))
        story.append(Spacer(1, 5))

    doc.build(story)
    return response


@tenant_required
@comptable_only
@transaction.atomic
def enregistrer_paiement_especes(request, eleve_id):
    eleve = Eleve.objects.get(pk=eleve_id)
    periode = None
    montant_restant = Decimal('0.00')

    if request.method == 'POST':
        form = PaiementForm(request.POST)
        if form.is_valid():
            paiement = form.save(commit=False)
            paiement.eleve = eleve

            # Récupération de la période de paiement
            periode = paiement.periode
            montant_total = periode.montant

            # Vérification du mode de paiement
            if periode.mode_paiement == 'FULL':
                # Validation paiement complet
                if paiement.montant_paye != montant_total:
                    form.add_error('montant_paye',
                                   f"Le paiement complet doit être de {montant_total} FCFA")
                    return render(request, 'paiement.html', {'form': form})

            else:  # Mode échelonné
                # Calcul du total déjà payé
                total_paye = Paiement.objects.filter(
                    periode=periode,
                    eleve=eleve,
                    statut_paiement='REUSSI'
                ).aggregate(total=models.Sum('montant_paye'))['total'] or Decimal('0.00')

                montant_restant = montant_total - total_paye

                # Validation du montant
                if paiement.montant_paye <= Decimal('0.00'):
                    form.add_error('montant_paye',
                                   "Le montant doit être positif")
                elif paiement.montant_paye > montant_restant:
                    form.add_error('montant_paye',
                                   f"Montant restant à payer : {montant_restant} FCFA")

            if not form.errors:
                # Marquer le paiement selon le solde
                if paiement.montant_paye == montant_restant:
                    paiement.statut_paiement = 'REUSSI'
                    messages.success(request, "Paiement final effectué avec succès !")
                else:
                    paiement.statut_paiement = 'PARTIEL'
                    messages.warning(request,
                                     f"Paiement partiel enregistré. Solde restant : {montant_restant - paiement.montant_paye} FCFA")

                paiement.save()

                # Mettre à jour le statut de suspension
                nouveau_solde = montant_restant - paiement.montant_paye
                if nouveau_solde <= Decimal('0.00'):
                    eleve.suspendu = False
                    eleve.save()
                    # verifier_suspension_eleve.delay(eleve.id)

                # Envoyer confirmation
                # envoyer_confirmation_paiement.delay(
                #     eleve.id,
                #     paiement.montant_paye,
                #     nouveau_solde,
                #     periode.nom
                # )

                return redirect('historique_paiements', eleve_id=eleve.id)

    else:
        form = PaiementForm()

    context = {
        'form': form,
        'eleve': eleve,
        'periode': periode,
        'montant_restant': montant_restant,
        'historique_paiements': Paiement.objects.filter(eleve=eleve).order_by('-date_paiement')
    }
    return render(request, 'APP_G2S/paiement/enregistrement.html', context)


@tenant_required
@comptable_only
def exporter_paiements_excel(request):
    paiements = Paiement.objects.select_related('eleve').filter(mode_paiement='ESPECES')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="historique_paiements.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Paiements en espèces"

    # En-têtes
    columns = [
        'Date', 'Référence', 'Élève',
        'Classe', 'Montant', 'Encaisseur'
    ]
    ws.append(columns)

    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Données
    for p in paiements:
        ws.append([
            p.date_paiement.strftime('%d/%m/%Y %H:%M'),
            p.reference,
            p.eleve.nom,
            p.eleve.prenom,
            p.eleve.nom_pere,
            p.eleve.prenom_pere,
            p.eleve.classe.nom,
            p.montant_paye,
            p.encaisser_par.get_full_name() if p.encaisser_par else 'Système'
        ])

    wb.save(response)
    return response

@tenant_required
@comptable_only
def recherche_eleves(request):
    query = request.GET.get('q', '')
    eleves = Eleve.objects.filter(
        models.Q(nom__icontains=query) |
        models.Q(prenom__icontains=query) |
        models.Q(telephone__icontains=query)
    ).select_related('classe')[:10]

    results = [{
        'id': e.id,
        'nom_complet': f"{e.prenom} {e.nom}",
        'classe': str(e.classe),
        'solde': e.solde_restant,
        'url': reverse('enregistrer_paiement_especes', args=[e.id])
    } for e in eleves]

    return JsonResponse({'results': results})

@tenant_required
@comptable_only
def annuler_paiement(request, paiement_id):
    paiement = get_object_or_404(Paiement, id=paiement_id)
    if paiement.mode_paiement == 'ESPECES':
        paiement.delete()
        messages.success(request, "Paiement en espèces annulé avec succès")
    return redirect('gerer_paiements')

@method_decorator(comptable_only, name='dispatch')
class DashboardPaiementsView(TemplateView):
    template_name = 'APP_G2S/composant-comptable/dashboard_paiements.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stats'] = {
            'paiements_mois': Paiement.objects.filter(
                date_paiement__month=timezone.now().month
            ).aggregate(Sum('montant_paye')),
            'evolution': Paiement.objects.annotate(
                mois=TruncMonth('date_paiement')
            ).values('mois').annotate(total=Sum('montant_paye'))
        }
        return context




@require_GET
@tenant_required
@administrateur_required
def get_emplois(request):
    eleve_id = request.GET.get('eleve_id')
    try:
        eleve = Eleve.objects.get(id=eleve_id)
        emplois = EmploiDuTemps.objects.filter(classe=eleve.classe).values('id', 'date', 'matiere__nom')
        return JsonResponse(list(emplois), safe=False)
    except Eleve.DoesNotExist:
        return JsonResponse([], safe=False)

@csrf_exempt
def callback_paiement(request):
    # Logique de vérification des transactions Orange/Malitel
    if request.method == 'POST':
        data = json.loads(request.body)
        # Valider la transaction avec l'API mobile money
        # Mettre à jour le statut du paiement
        return JsonResponse({'status': 'success'})
    return None


@tenant_required
@administrateur_required
def exporter_excel_absences(request):
    absences = Absence.objects.select_related('eleve', 'emploi_du_temps').order_by('-date')
    print(absences)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="absences"du_{timezone.now()}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Absences"

    columns = ['Élève', 'Date', 'Matière', 'Statut', 'Justificatif']
    ws.append(columns)

    for absence in absences:
        ws.append([
            f"{absence.eleve.prenom} {absence.eleve.nom}",
            absence.date,
            absence.emploi_du_temps.matiere.nom,
            absence.get_justification_status_display(),
            "Oui" if absence.justification_document else "Non"
        ])

    wb.save(response)
    return response



@method_decorator(comptable_only, name='dispatch')
class ChoixModePaiementView(TemplateView):
    template_name = 'APP_G2S/composant-comptable/choix_paiement.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        eleve_id = self.kwargs.get('eleve_id')
        eleve = get_object_or_404(Eleve, id=eleve_id)
        periodes_impayees = PeriodePaiement.objects.filter(
            classe=eleve.classe,
            date_fin__lt=datetime.today()
        ).exclude(paiements__eleve=eleve)

        context['eleve'] = eleve
        context['periodes_impayees'] = periodes_impayees
        return context

@method_decorator(csrf_exempt, name='dispatch')
class InitierPaiementMobileView(View):

    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        eleve_id = data['eleve_id']
        montant = data['montant']
        telephone = data['telephone']
        operator = data['operator']
        eleve = get_object_or_404(Eleve, id=eleve_id)

        try:
            if operator == 'ORANGE':
                response = OrangeMoneyAPI.initier_paiement(montant, telephone)
            elif operator == 'MALITEL':
                response = MalitelMoneyAPI.initier_paiement(montant, telephone)
            else:
                return JsonResponse({'status': 'error', 'message': 'Opérateur invalide'})

            # Enregistrer la transaction en base
            Paiement.objects.create(
                eleve=eleve,
                periode_id=data['periode_id'],
                montant_paye=data['montant'],
                mode_paiement=data['operator'],
                statut_paiement='EN_ATTENTE'
            )
            return JsonResponse(response)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    def get(self, request, *args, **kwargs):
        return HttpResponseForbidden()


@method_decorator(multi_role_required(directeur_required, comptable_required), name='dispatch')
class HistoriquePaiementsView(TemplateView):
    template_name = 'APP_G2S/composant-comptable/historique_paiement.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        eleve_id = self.kwargs.get('eleve_id')
        paiements = Paiement.objects.select_related('eleve', 'periode').order_by('-date')
        eleve = get_object_or_404(Eleve, id=eleve_id)

        context['paiements'] = paiements
        context['eleve'] = eleve
        return context






'''

la partie dedie aux vues de modificarions


'''


@tenant_required
@censeur_only
@requires_approval(action_type='modifier_examen', model=Examen, id_param='examen_id')
def modifier_examen(request, examen_id):
    examen = get_object_or_404(Examen, id=examen_id)
    if request.method == 'POST':
        form = ExamenForm(request.POST, instance=examen)
        if form.is_valid():
            form.save()
            messages.success(request, "Examen modifié avec succès !")
            return redirect('liste_examens')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ExamenForm(instance=examen)

    return render(request, 'APP_G2S/composant-admin/modifier_examen.html', {
        'form': form,
        'examen': examen
    })


@tenant_required
@administrateur_required

def ajouter_historique_academique(request):
    if request.method == 'POST':
        form = HistoriqueAcademiqueForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Historique académique ajouté avec succès.")
            return redirect('liste_historique_academique')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = HistoriqueAcademiqueForm()
    return render(request, 'APP_G2S/composant-admin/ajouter_historique_academique.html', {'form': form})

@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required, surveillant_required)
def liste_historique_academique(request):
    historiques = HistoriqueAcademique.objects.select_related('eleve', 'periode').order_by('-periode')
    return render(request, 'APP_G2S/composant-admin/liste_historique_academique.html', {'historiques': historiques})

@tenant_required
@administrateur_required
@multi_role_required(directeur_required, censeur_required)
def statuts_academiques(request):
    # Récupérer la dernière période clôturée
    derniere_periode = Periode.objects.filter(cloture=True).order_by('-date_fin').first()
    # Si aucune période clôturée, utiliser la période active
    periode = derniere_periode or Periode.objects.filter(is_active=True).first()

    # Subquerys pour dernier statut et moyenne
    dernier_statut_subquery = Subquery(
        HistoriqueAcademique.objects.filter(
            eleve=OuterRef('pk')
        ).order_by('-periode__date_fin').values('decision')[:1]
    )
    derniere_moyenne_subquery = Subquery(
        HistoriqueAcademique.objects.filter(
            eleve=OuterRef('pk')
        ).order_by('-periode__date_fin').values('moyenne')[:1]
    )

    eleves = Eleve.objects.annotate(
        dernier_statut=dernier_statut_subquery,
        derniere_moyenne=derniere_moyenne_subquery
    ).select_related('classe')

    # Catégoriser les élèves
    promus = []
    redoublants = []
    expulses = []

    for eleve in eleves:
        if eleve.est_expulse:
            expulses.append(eleve)
        elif eleve.dernier_statut == 'ADMIS' and getattr(eleve.classe, 'niveau_superieur', None):
            promus.append(eleve)
        elif eleve.dernier_statut == 'REDOUBLE':
            redoublants.append(eleve)

    context = {
        'promus': promus,
        'redoublants': redoublants,
        'expulses': expulses,
        'periode': periode,
        'total_eleves': eleves.count(),
        'now': timezone.now()
    }

    return render(request, 'APP_G2S/composant-admin/statuts_academiques.html', context)


def erreur_user_type(request):
    """
    Vue pour afficher une erreur de type d'utilisateur.
    Cette vue est appelée lorsqu'un utilisateur tente d'accéder à une fonctionnalité
    avec un type d'utilisateur incompatible.
    """
    error_message = request.GET.get('message', 'Type d\'utilisateur incompatible avec cette action.')
    return render(request, 'error_flotant/erreur_user_type.html', {'error_message': error_message})
